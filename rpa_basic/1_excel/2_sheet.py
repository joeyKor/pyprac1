from openpyxl import Workbook
wb= Workbook()

# wb.active

ws = wb.create_sheet()
ws.title = "MySheet" # Sheet 이름 병경
ws.sheet_properties.tabColor = "ff66ff"


ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet" , 2 )

wb.save("practice.xlsx")