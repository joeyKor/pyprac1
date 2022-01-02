from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

ws.append(["번호" , "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0,100), randint(0,100)])
 

col_B = ws["B"]

for cell in col_B:
    print(cell.value)
    
col_range = ws["B:C"]
#for cols in col_range:
#    for cell in cols:
#        print(cell.value)

row_title = ws[1] 
for cell in row_title:
    print(cell.value)


row_range = ws[2:6] # 2row to 6row load
for rows in row_range:
    for cell in rows:
        print(cell.value, end="  ")
    print()

wb.save("sample.xlsx")