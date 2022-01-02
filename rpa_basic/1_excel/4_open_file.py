from openpyxl import load_workbook # load file
wb = load_workbook("sample.xlsx")
ws = wb.active

# load cell data

for x in range(1,11):
    for y in range(1, 11):
        print(ws.cell(row=x, column =y).value, end="  ")
    print()

# if you don know cell
for x in range(1, ws.max_row +1):
 for  y in range(1, ws.max_column +1):
     print(ws.cell(row=x, column= y).value, end="  ")
 print()