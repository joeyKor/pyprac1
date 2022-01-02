from openpyxl import Workbook
wb= Workbook()
ws = wb.active
ws.title = "NadoSheet"


ws["A1"] =2
ws["A2"] =2
ws["A3"] =3

ws["B1"] =4
ws["B2"] =5
ws["B3"] =6

c = ws.cell(column=3, row=1, value="this is sparta")
print(c.value)

from random import *

index = 1

for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column = y, value=index)
        index += 1

print(ws["A1"])
print(ws["A1"].value)

wb.save("sample.xlsx")