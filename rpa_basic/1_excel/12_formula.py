import datetime

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["a1"] = datetime.datetime.today()

ws["a2"] = "=sum(1,2,3)"
ws["a3"] = "=average(1,2,3)"

wb.save("sample_formual.xlsx")