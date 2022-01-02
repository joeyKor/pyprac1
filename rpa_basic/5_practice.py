from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["num", "kor", "eng"])

for i in range(1,10):
    ws.append([i,randint(50,100), randint(50,100)])

col_B = ws["B"]

# for cell in col_B:
#     print(cell.value)

# row_range = ws[2:6]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end ="  ")
#     print()

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end = " ")

#     print()

# for row in ws.iter_rows():
#     print(row[2].value)

for column in ws.iter_cols():
    print(column[0].value)



wb.save("samplemade.xlsx")