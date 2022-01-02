from openpyxl import load_workbook
wb= load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # num, eng, math
    if int(row[1].value) > 80:
        print(row[0].value,"번 학생은 영어 지니어스")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "eng":
            cell.value = "com"
            print("this is modified")

wb.save("sampel_modified.xlsx")
