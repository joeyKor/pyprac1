from openpyxl import Workbook
wb = Workbook()
ws =wb.create_sheet()
ws1 = wb.create_sheet("This Sheet")

new_ws = wb["This Sheet"]

print(wb.sheetnames)

ws.title = "MySheet"
new_ws["A1"] = "Test"

target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample21.xlsx")