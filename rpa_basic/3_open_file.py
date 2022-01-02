from openpyxl import load_workbook
wb= load_workbook("sample.xlsx")
ws = wb.active

ws.move_range("B1:c11", rows =0, cols =1)

ws["B1"].value = "kor"

wb.save("sample_kor.xlsx")